"""
Справочник клиентов из HelpDeskEddy (xlsx «клиенты из эдди для комблока»).

Назначение: расшифровать в обращении кодовые имена owner'ов, названия клиентов
и сокращения, которых нет в базе знаний. Это НЕ семантический поиск — таблица
маленькая (сотни строк) и матчится точным вхождением ключей (owner-слаг, название
компании, домен почты, главный заказчик) в текст тикета. Найденные строки
подставляются в контекст LLM отдельным блоком и помогают погасить unknown_terms.

Колонки файла (заголовки распознаются по подстроке, регистр не важен):
    компания | главный заказчик | почта | категория | договор | платформа |
    овнер | активность пу… | количество пу | количество бс | есть в биллинге |
    комментарий

Обновление справочника = заменить xlsx и перезапустить процесс (грузится один раз).
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from functools import lru_cache
from pathlib import Path

from config import config

log = logging.getLogger(__name__)

# Канонические поля → подстроки, по которым ищем заголовок колонки в файле.
_COLUMN_ALIASES = {
    "company": ("компан",),
    "customer": ("заказчик",),
    "email": ("почт", "mail", "e-mail"),
    "category": ("категор",),
    "contract": ("договор",),
    "platform": ("платформ",),
    "owner": ("овнер", "owner"),
    "comment": ("коммент",),
}

# Колонки, значения которых служат ключами поиска в тексте тикета.
_KEY_FIELDS = ("owner", "company", "customer", "email")
# Колонки, попадающие в подставляемую LLM карточку клиента (в этом порядке).
_DISPLAY_FIELDS = (
    ("company", "клиент"),
    ("owner", "owner"),
    ("category", "категория"),
    ("contract", "договор"),
    ("platform", "платформа"),
    ("comment", "комментарий"),
)

# Заглушки, которые в файле означают «значения нет» — в ключи/карточку не идут.
_BLANKS = {"", "нет данных", "-", "—", "н/д"}
# Минимальная длина ключа: короче — слишком шумно (ложные срабатывания).
_MIN_NEEDLE = 4
_MAX_COMMENT = 200


def _norm(v: object) -> str:
    return str(v if v is not None else "").strip()


@lru_cache(maxsize=8192)
def _matcher(needle: str):
    """Скомпилировать проверку вхождения ключа в текст (уже в нижнем регистре).

    Ключ из «словных» символов (буквы/цифры/_) матчим по границам слова, чтобы
    'altke' не сработал внутри другого слова. Ключ со спецсимволами (домен,
    e-mail) — обычным вхождением: он и так достаточно специфичен.
    """
    if re.fullmatch(r"\w+", needle, re.UNICODE):
        pat = re.compile(rf"\b{re.escape(needle)}\b", re.UNICODE)
        return lambda hay: pat.search(hay) is not None
    return lambda hay: needle in hay


@dataclass
class ClientRow:
    company: str = ""
    customer: str = ""
    email: str = ""
    category: str = ""
    contract: str = ""
    platform: str = ""
    owner: str = ""
    comment: str = ""
    needles: tuple[str, ...] = field(default_factory=tuple)

    def matches(self, text_lower: str) -> bool:
        return any(_matcher(n)(text_lower) for n in self.needles)

    def as_card(self, index: int) -> str:
        parts = []
        for fname, label in _DISPLAY_FIELDS:
            val = getattr(self, fname)
            if fname == "comment":
                val = val[:_MAX_COMMENT]
            if val and val.lower() not in _BLANKS:
                parts.append(f"{label}: {val}")
        return f"[{index}] " + " | ".join(parts)


def _build_needles(row: ClientRow) -> tuple[str, ...]:
    needles: set[str] = set()
    for fname in _KEY_FIELDS:
        val = _norm(getattr(row, fname)).lower()
        if not val or val in _BLANKS:
            continue
        if fname == "email" and "@" in val:
            val = val.split("@", 1)[1]  # домен почты — он идентифицирует клиента
        if fname == "owner" and "@" in val:
            needles.add(val.split("@", 1)[0])  # слаг до @ (aurora@lar.cloud → aurora)
        if len(val) >= _MIN_NEEDLE:
            needles.add(val)
    return tuple(sorted(needles))


class Directory:
    """Справочник клиентов: матчинг строк по тексту тикета."""

    def __init__(self, rows: list[ClientRow]) -> None:
        self.rows = rows

    def __len__(self) -> int:
        return len(self.rows)

    def match(self, text: str, limit: int = 8) -> list[ClientRow]:
        low = (text or "").lower()
        if not low:
            return []
        out: list[ClientRow] = []
        for row in self.rows:
            if row.matches(low):
                out.append(row)
                if len(out) >= limit:
                    break
        return out

    def format_block(self, rows: list[ClientRow]) -> str:
        """Блок-карточки совпавших клиентов для контекста LLM ('' — если пусто)."""
        if not rows:
            return ""
        lines = ["# Справочник клиентов (совпадения по обращению)"]
        lines += [r.as_card(i) for i, r in enumerate(rows, 1)]
        return "\n".join(lines)


def _resolve_columns(header: list[str]) -> dict[str, int]:
    """Сопоставить канонические поля с индексами колонок по подстроке заголовка."""
    mapping: dict[str, int] = {}
    for idx, raw in enumerate(header):
        name = _norm(raw).lower()
        if not name:
            continue
        for field_name, aliases in _COLUMN_ALIASES.items():
            if field_name in mapping:
                continue
            if any(a in name for a in aliases):
                mapping[field_name] = idx
    return mapping


def load_directory(path: str | None = None) -> Directory:
    """Прочитать xlsx-справочник. Файла нет/пустой/без ключевых колонок → пустой
    справочник (мягкая деградация: пайплайн работает и без него)."""
    target = Path(path or config.rag.directory_path).expanduser()
    if not target.exists():
        log.info("справочник клиентов не найден: %s — работаю без него", target)
        return Directory([])

    from openpyxl import load_workbook

    wb = load_workbook(target, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows_iter))
    except StopIteration:
        wb.close()
        return Directory([])

    cols = _resolve_columns([_norm(h) for h in header])
    if "owner" not in cols and "company" not in cols:
        log.warning("справочник %s: не нашёл колонок owner/компания — пропускаю", target)
        wb.close()
        return Directory([])

    def cell(values, fname):
        idx = cols.get(fname)
        return _norm(values[idx]) if idx is not None and idx < len(values) else ""

    rows: list[ClientRow] = []
    for values in rows_iter:
        if values is None or not any(_norm(v) for v in values):
            continue
        row = ClientRow(
            company=cell(values, "company"),
            customer=cell(values, "customer"),
            email=cell(values, "email"),
            category=cell(values, "category"),
            contract=cell(values, "contract"),
            platform=cell(values, "platform"),
            owner=cell(values, "owner"),
            comment=cell(values, "comment"),
        )
        row.needles = _build_needles(row)
        if row.needles:  # без ключей строку не сматчить — не храним
            rows.append(row)
    wb.close()
    log.info("справочник клиентов: %d строк из %s", len(rows), target)
    return Directory(rows)


_directory: Directory | None = None


def get_directory() -> Directory:
    """Лениво загруженный singleton справочника (как пайплайн в processing)."""
    global _directory
    if _directory is None:
        _directory = load_directory()
    return _directory
