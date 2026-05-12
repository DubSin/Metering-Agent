"""
Генерация артефактов для клиента: xlsx-отчёт + опц. Playwright-скриншот
страницы «Журналы» Metering Server.
"""
from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

from langchain_core.tools import tool
from openpyxl import Workbook
from openpyxl.styles import Font

from config import config

log = logging.getLogger(__name__)


def _reports_dir() -> Path:
    p = Path(config.reports_dir)
    p.mkdir(parents=True, exist_ok=True)
    return p


def _ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


# ---------- xlsx ----------

@tool
async def build_readings_report(
    task_id: str,
    readings: dict,
    kind: str = "daily",
) -> str:
    """Собрать xlsx-отчёт с показаниями.

    readings: {meter_id: [{timestamp, register, value, unit}, ...]}
    kind: daily | last | collection_map — влияет на название листа.
    Возвращает абсолютный путь к файлу.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = {
        "daily": "Суточные",
        "last": "Последние показания",
        "collection_map": "Карта сбора",
    }.get(kind, "Показания")

    headers = ["ПУ", "Дата/время", "Регистр", "Значение", "Ед."]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)

    for meter_id, rows in (readings or {}).items():
        if not rows:
            ws.append([meter_id, "—", "—", "нет данных", ""])
            continue
        for row in rows:
            ws.append([
                meter_id,
                row.get("timestamp", ""),
                row.get("register", ""),
                row.get("value", ""),
                row.get("unit", ""),
            ])

    for col in ws.columns:
        width = max(len(str(c.value or "")) for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(width, 40)

    path = _reports_dir() / f"readings_{task_id}_{kind}_{_ts()}.xlsx"
    wb.save(path)
    log.info("readings report saved: %s", path)
    return str(path.resolve())


@tool
async def build_status_report(task_id: str, statuses: list[dict]) -> str:
    """Собрать xlsx со статусами вывода ПУ на связь.

    statuses: [{meter_id, online: bool|None, detail, last_seen?}, ...]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Статусы ПУ"

    ws.append(["ПУ", "На связи", "Подробности", "Последняя связь"])
    for c in ws[1]:
        c.font = Font(bold=True)

    for s in statuses or []:
        online = s.get("online")
        ws.append([
            s.get("meter_id", ""),
            "да" if online is True else ("нет" if online is False else "—"),
            s.get("detail", ""),
            s.get("last_seen", ""),
        ])

    for col in ws.columns:
        width = max(len(str(c.value or "")) for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(width, 40)

    path = _reports_dir() / f"status_{task_id}_{_ts()}.xlsx"
    wb.save(path)
    log.info("status report saved: %s", path)
    return str(path.resolve())


# ---------- Playwright-скриншот ----------

@tool
async def capture_metering_screenshot(
    task_id: str,
    page_url: str,
    selector: str | None = None,
) -> str:
    """Снять скриншот страницы Metering Server (например, «Журналы»).

    page_url — абсолютный URL страницы в UI Metering Server.
    selector — опц. CSS-селектор конкретного блока (если нужен фрагмент, а не вся страница).
    Требуется установленный playwright (`playwright install chromium`).
    """
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        return "Playwright не установлен: pip install playwright && playwright install chromium"

    out = _reports_dir() / f"screenshot_{task_id}_{_ts()}.png"

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        try:
            ctx_kwargs: dict = {}
            if config.metering.username:
                ctx_kwargs["http_credentials"] = {
                    "username": config.metering.username,
                    "password": config.metering.password,
                }
            ctx = await browser.new_context(**ctx_kwargs)
            page = await ctx.new_page()
            await page.goto(page_url, wait_until="networkidle")
            if selector:
                el = await page.query_selector(selector)
                if el is None:
                    await page.screenshot(path=str(out), full_page=True)
                else:
                    await el.screenshot(path=str(out))
            else:
                await page.screenshot(path=str(out), full_page=True)
        finally:
            await browser.close()

    log.info("screenshot saved: %s", out)
    return str(out.resolve())
